import sys
import os
import json
import time
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QPushButton, QTextBrowser, QFileDialog, QTableWidget,
    QTableWidgetItem, QHeaderView, QMessageBox, QGroupBox, QFormLayout, QCheckBox,
    QDialog, QDialogButtonBox, QTextEdit, QSpinBox
)
from PySide6.QtCore import QThread, Signal, QObject

from scraper.universal_scraper import UniversalFacultyScraper, Faculty
from parser.folder_loader import FolderCampaignLoader
from mailer.gmail_sender import GmailSender
from ai_personalizer import AIPersonalizer
from recipient_filter import filter_recipients, format_preview

SETTINGS_FILE = "settings.json"
SEND_HISTORY_FILE = "send_history.json"
HISTORY_RETENTION_DAYS = 30
DAY_SECONDS = 24 * 60 * 60


def _load_send_history():
    try:
        with open(SEND_HISTORY_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, list):
            cutoff = time.time() - HISTORY_RETENTION_DAYS * DAY_SECONDS
            return [h for h in data
                    if isinstance(h, dict) and float(h.get("ts", 0)) >= cutoff]
    except Exception:
        pass
    return []


def _save_send_history(history):
    try:
        with open(SEND_HISTORY_FILE, "w", encoding="utf-8") as f:
            json.dump(history, f)
    except Exception:
        pass


def _sent_last_24h(history):
    cutoff = time.time() - DAY_SECONDS
    return sum(1 for h in history if float(h.get("ts", 0)) >= cutoff)


def _emails_last_24h(history):
    cutoff = time.time() - DAY_SECONDS
    return {(h.get("email", "") or "").strip().lower()
            for h in history if float(h.get("ts", 0)) >= cutoff}


# --- Auto-send scheduling window + helpers --------------------------------
PENDING_QUEUE_FILE = "pending_queue.json"
ALLOWED_START_HOUR = 8     # 8 AM  - never auto-send before this
ALLOWED_END_HOUR = 21      # 9 PM  - never auto-send after this


def _in_allowed_hours(dt=None):
    dt = dt or datetime.now()
    return ALLOWED_START_HOUR <= dt.hour < ALLOWED_END_HOUR


def _next_allowed_start(dt):
    start = dt.replace(hour=ALLOWED_START_HOUR, minute=0, second=0, microsecond=0)
    end = dt.replace(hour=ALLOWED_END_HOUR, minute=0, second=0, microsecond=0)
    if dt < start:
        return start
    if dt >= end:
        return (dt + timedelta(days=1)).replace(
            hour=ALLOWED_START_HOUR, minute=0, second=0, microsecond=0)
    return dt


def _headroom_available_epoch(history, cap, need, now=None):
    """Epoch time when there will be room to send `need` more emails under
    `cap`, based on when past sends roll out of the rolling 24h window."""
    now = now or time.time()
    if not cap:
        return now
    in_window = sorted(float(h.get("ts", 0)) for h in history
                       if float(h.get("ts", 0)) >= now - DAY_SECONDS)
    headroom = cap - len(in_window)
    if headroom >= need:
        return now
    need_expire = need - headroom
    if need_expire > len(in_window):
        need_expire = len(in_window)
    if need_expire <= 0:
        return now
    return max(in_window[need_expire - 1] + DAY_SECONDS, now)


def _effective_resume_dt(history, cap, need):
    dt = datetime.fromtimestamp(_headroom_available_epoch(history, cap, need))
    if not _in_allowed_hours(dt):
        dt = _next_allowed_start(dt)
    return dt


def _format_clock(dt):
    now = datetime.now()
    h = dt.hour % 12 or 12
    ampm = "AM" if dt.hour < 12 else "PM"
    clock = f"{h}:{dt.minute:02d} {ampm}"
    if dt.date() == now.date():
        day = "today"
    elif dt.date() == (now + timedelta(days=1)).date():
        day = "tomorrow"
    else:
        day = dt.strftime("%a %b %d")
    return f"{clock} {day}"


def _faculty_to_dict(fac):
    return {
        "full_name": getattr(fac, "full_name", ""),
        "first_name": getattr(fac, "first_name", ""),
        "last_name": getattr(fac, "last_name", ""),
        "title": getattr(fac, "title", ""),
        "email": getattr(fac, "email", ""),
        "profile_url": getattr(fac, "profile_url", ""),
        "bio_text": getattr(fac, "bio_text", "") or getattr(fac, "title", ""),
    }


def _faculty_from_dict(d):
    return Faculty(
        full_name=d.get("full_name", "") or f"{d.get('first_name','')} {d.get('last_name','')}".strip(),
        first_name=d.get("first_name", ""),
        last_name=d.get("last_name", ""),
        title=d.get("title", ""),
        email=(d.get("email", "") or "").lower(),
        profile_url=d.get("profile_url", ""),
        email_source="pending",
        bio_text=d.get("bio_text", ""),
    )


def _save_pending_queue(targets, source_label, university):
    try:
        if not targets:
            if os.path.exists(PENDING_QUEUE_FILE):
                os.remove(PENDING_QUEUE_FILE)
            return
        data = {
            "source_label": source_label,
            "university": university,
            "saved_ts": time.time(),
            "targets": [_faculty_to_dict(f) for f in targets],
        }
        with open(PENDING_QUEUE_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f)
    except Exception:
        pass


def _load_pending_queue():
    try:
        with open(PENDING_QUEUE_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        targets = [_faculty_from_dict(d) for d in data.get("targets", [])]
        return targets, data.get("source_label", "saved queue"), data.get("university", "")
    except Exception:
        return [], "", ""


def confirm_send_dialog(parent, header, body_text):
    """
    A confirmation dialog with a SCROLLABLE text area (unlike QMessageBox,
    which can't scroll and runs off-screen for long recipient lists).
    Returns True if the user clicks Yes, False otherwise.
    """
    dlg = QDialog(parent)
    dlg.setWindowTitle("Confirm send")
    dlg.resize(640, 560)
    layout = QVBoxLayout(dlg)

    label = QLabel(header)
    label.setWordWrap(True)
    layout.addWidget(label)

    text = QTextEdit()
    text.setReadOnly(True)
    text.setLineWrapMode(QTextEdit.NoWrap)   # keep columns aligned; horizontal scroll if needed
    text.setPlainText(body_text)
    layout.addWidget(text)                    # expands and scrolls

    buttons = QDialogButtonBox(QDialogButtonBox.Yes | QDialogButtonBox.No)
    buttons.button(QDialogButtonBox.No).setDefault(True)   # default to No
    buttons.accepted.connect(dlg.accept)
    buttons.rejected.connect(dlg.reject)
    layout.addWidget(buttons)

    return dlg.exec() == QDialog.Accepted


class ScrapeWorker(QObject):
    finished = Signal(list)
    progress = Signal(str)

    def __init__(self, url, university_name=""):
        super().__init__()
        self.url = url
        self.university_name = university_name

    def run(self):
        scraper = UniversalFacultyScraper()
        results = scraper.scrape(self.url, university_name=self.university_name,
                                 progress_cb=self.progress.emit)
        self.finished.emit(results)


class SendWorker(QObject):
    """Sends one batch of emails on a background thread so the UI never freezes.
    Emits a signal per send so the main thread can record history, update the
    24h counter, and shrink the persisted pending queue as it goes."""
    progress = Signal(str)
    sent_one = Signal(str)          # email
    failed_one = Signal(str, str)   # email, error
    finished = Signal(int, int)     # sent_count, fail_count

    def __init__(self, targets, campaign_data, ai_personalizer, university,
                 temp_dir, delay, gmail, pwd):
        super().__init__()
        self.targets = targets
        self.campaign_data = campaign_data
        self.ai = ai_personalizer
        self.university = university
        self.temp_dir = temp_dir
        self.delay = delay
        self.gmail = gmail
        self.pwd = pwd
        self._stop = False

    def stop(self):
        self._stop = True

    def run(self):
        sent = 0
        failed = 0
        total = len(self.targets)
        try:
            with GmailSender(self.gmail, self.pwd, delay=self.delay) as sender:
                for i, fac in enumerate(self.targets, 1):
                    if self._stop:
                        self.progress.emit("Stopped by user.")
                        break

                    profile_hook = ""
                    if self.ai and getattr(fac, "profile_url", ""):
                        try:
                            profile_hook = self.ai.generate_custom_hook(
                                professor_name=fac.full_name,
                                bio_text=getattr(fac, "bio_text", "") or fac.title,
                            )
                        except Exception:
                            profile_hook = ""

                    subject = FolderCampaignLoader.personalize_text(
                        self.campaign_data["subject"], fac.first_name, fac.last_name,
                        fac.title, university=self.university, profile_match=profile_hook,
                    )
                    html_body = FolderCampaignLoader.personalize_text(
                        self.campaign_data["html_body"], fac.first_name, fac.last_name,
                        fac.title, university=self.university, profile_match=profile_hook,
                    )

                    attachment_cv = self.campaign_data["cv_path"]
                    if self.campaign_data["cv_type"] == "docx":
                        output_pdf = os.path.join(self.temp_dir, f"CV_{fac.last_name or 'Faculty'}.pdf")
                        replacements = {
                            "{PROFESSOR_NAME}": fac.full_name,
                            "{LAST_NAME}": fac.last_name,
                            "{UNIVERSITY}": self.university,
                            "{PROFILE_MATCH}": profile_hook,
                        }
                        try:
                            attachment_cv = FolderCampaignLoader.generate_tailored_cv(
                                self.campaign_data["cv_path"], output_pdf, replacements
                            )
                        except Exception as e:
                            failed += 1
                            self.failed_one.emit(fac.email, f"CV generation failed: {e}")
                            continue

                    try:
                        sender.send_email(
                            fac.email, subject, html_body,
                            attachment_path=attachment_cv,
                            attachment_name="CV_FAC.pdf",
                            inline_images=self.campaign_data.get("inline_images"),
                        )
                        sent += 1
                        self.sent_one.emit(fac.email)
                        self.progress.emit(f"Sent {i}/{total}: {fac.email}")
                    except Exception as e:
                        failed += 1
                        self.failed_one.emit(fac.email, str(e))
                        self.progress.emit(f"Failed {i}/{total}: {fac.email}")
        except Exception as e:
            self.progress.emit(f"Connection error: {e}")
        self.finished.emit(sent, failed)


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("FacultyMailer - Universal Outreach Engine")
        self.resize(1100, 900)
        self.faculty_data = []
        self.campaign_data = None
        self.send_history = _load_send_history()
        # Auto-send / pause-resume state
        self.pending_targets, self._pending_source, self._pending_uni = _load_pending_queue()
        self._sending = False
        self._scanning = False
        self._send_thread = None
        self._send_worker = None
        self._run_failures = []
        self._run_interactive = False

        self.init_ui()
        self.load_settings()

    def init_ui(self):
        main_layout = QVBoxLayout()

        config_box = QGroupBox("Campaign & Directory Inputs")
        config_layout = QFormLayout()

        self.uni_input = QLineEdit()
        self.uni_input.setPlaceholderText("e.g., Carnegie Mellon University")
        self.uni_input.textChanged.connect(self.save_settings)

        self.url_input = QLineEdit()
        self.url_input.setPlaceholderText("e.g., https://csd.cmu.edu/people/faculty")
        self.url_input.textChanged.connect(self.save_settings)

        folder_sub_layout = QHBoxLayout()
        self.folder_input = QLineEdit()
        self.folder_input.setReadOnly(True)
        self.folder_input.setPlaceholderText("Path to folder with Subject.txt, Email.docx, CV_FAC.docx/pdf")
        self.folder_btn = QPushButton("Browse Folder")
        self.folder_btn.clicked.connect(self.browse_campaign_folder)
        folder_sub_layout.addWidget(self.folder_input)
        folder_sub_layout.addWidget(self.folder_btn)

        excel_sub_layout = QHBoxLayout()
        self.excel_input = QLineEdit()
        self.excel_input.setPlaceholderText("Path to save output .xlsx file")
        self.excel_btn = QPushButton("Save Excel To...")
        self.excel_btn.clicked.connect(self.browse_excel_output)
        self.excel_input.textChanged.connect(self.save_settings)
        excel_sub_layout.addWidget(self.excel_input)
        excel_sub_layout.addWidget(self.excel_btn)

        ai_sub_layout = QHBoxLayout()
        self.ai_checkbox = QCheckBox("Enable AI Profile Hook ({PROFILE_MATCH})")
        self.ai_checkbox.stateChanged.connect(self.save_settings)
        self.gemini_key_input = QLineEdit()
        self.gemini_key_input.setEchoMode(QLineEdit.Password)
        self.gemini_key_input.setPlaceholderText("Gemini API Key (for AI profile matching)")
        self.gemini_key_input.textChanged.connect(self.save_settings)
        ai_sub_layout.addWidget(self.ai_checkbox)
        ai_sub_layout.addWidget(self.gemini_key_input)

        config_layout.addRow("University Name:", self.uni_input)
        config_layout.addRow("Faculty Directory URL:", self.url_input)
        config_layout.addRow("Campaign Folder:", folder_sub_layout)
        config_layout.addRow("Output Excel File (.xlsx):", excel_sub_layout)
        config_layout.addRow("AI Personalization:", ai_sub_layout)

        config_box.setLayout(config_layout)
        main_layout.addWidget(config_box)

        action_layout = QHBoxLayout()
        self.scan_btn = QPushButton("FIND FACULTY & CRAWL")
        self.scan_btn.setStyleSheet("font-weight: bold; padding: 6px;")
        self.scan_btn.clicked.connect(self.start_scan)

        self.reset_btn = QPushButton("Reset Saved Settings")
        self.reset_btn.clicked.connect(self.reset_settings)

        action_layout.addWidget(self.scan_btn)
        action_layout.addWidget(self.reset_btn)
        main_layout.addLayout(action_layout)

        self.status_label = QLabel("Status: Ready")
        main_layout.addWidget(self.status_label)

        self.table = QTableWidget(0, 4)
        self.table.setHorizontalHeaderLabels(["Name", "Title", "Email", "Status"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        main_layout.addWidget(self.table)

        preview_box = QGroupBox("Campaign Email Preview")
        preview_layout = QVBoxLayout()

        self.subject_preview = QLineEdit()
        self.subject_preview.setReadOnly(True)

        self.body_preview = QTextBrowser()

        preview_layout.addWidget(QLabel("Subject Preview:"))
        preview_layout.addWidget(self.subject_preview)
        preview_layout.addWidget(QLabel("Body HTML Preview (Use {PROFILE_MATCH} in Email.docx for AI sentence):"))
        preview_layout.addWidget(self.body_preview)
        preview_box.setLayout(preview_layout)
        main_layout.addWidget(preview_box)

        cred_box = QGroupBox("Gmail Dispatch Settings")
        cred_layout = QVBoxLayout()

        # Row 1: credentials
        creds_row = QHBoxLayout()
        self.gmail_input = QLineEdit()
        self.gmail_input.setPlaceholderText("Your Gmail Address")
        self.gmail_input.textChanged.connect(self.save_settings)

        self.pass_input = QLineEdit()
        self.pass_input.setEchoMode(QLineEdit.Password)
        self.pass_input.setPlaceholderText("Gmail App Password")
        self.pass_input.textChanged.connect(self.save_settings)
        creds_row.addWidget(self.gmail_input)
        creds_row.addWidget(self.pass_input)
        cred_layout.addLayout(creds_row)

        # Row 2: send from a selected Excel file
        excel_send_row = QHBoxLayout()
        self.send_excel_input = QLineEdit()
        self.send_excel_input.setReadOnly(True)
        self.send_excel_input.setPlaceholderText("Excel (.xlsx) to send from - the file your scraper saved")
        self.send_excel_input.textChanged.connect(self.save_settings)
        self.send_excel_btn = QPushButton("Choose Excel...")
        self.send_excel_btn.clicked.connect(self.browse_send_excel)
        self.send_from_excel_btn = QPushButton("SEND FROM EXCEL")
        self.send_from_excel_btn.setStyleSheet(
            "font-weight: bold; background-color: #107C41; color: white; padding: 8px;")
        self.send_from_excel_btn.clicked.connect(self.send_from_excel)
        excel_send_row.addWidget(QLabel("Send From Excel:"))
        excel_send_row.addWidget(self.send_excel_input)
        excel_send_row.addWidget(self.send_excel_btn)
        excel_send_row.addWidget(self.send_from_excel_btn)
        cred_layout.addLayout(excel_send_row)

        # Row 3: 24h cap + counter + send scraped list
        cap_row = QHBoxLayout()
        cap_row.addWidget(QLabel("Daily cap (last 24h):"))
        self.cap_input = QSpinBox()
        self.cap_input.setRange(0, 5000)      # 0 = no cap
        self.cap_input.setValue(200)
        self.cap_input.setToolTip("Maximum emails allowed per rolling 24 hours. 0 = no cap.")
        self.cap_input.valueChanged.connect(self._on_cap_changed)
        cap_row.addWidget(self.cap_input)

        self.sent_counter_label = QLabel("Sent last 24h: 0")
        self.sent_counter_label.setStyleSheet("font-weight: bold; padding: 0 12px;")
        cap_row.addWidget(self.sent_counter_label)

        self.refresh_counter_btn = QPushButton("Refresh")
        self.refresh_counter_btn.clicked.connect(self.update_sent_counter_label)
        cap_row.addWidget(self.refresh_counter_btn)

        cap_row.addStretch(1)

        self.send_btn = QPushButton("SEND ALL EMAILS")
        self.send_btn.setStyleSheet(
            "font-weight: bold; background-color: #0078D4; color: white; padding: 8px;")
        self.send_btn.clicked.connect(self.send_emails)
        cap_row.addWidget(self.send_btn)
        cred_layout.addLayout(cap_row)

        # Row 4: auto-send status + controls
        auto_row = QHBoxLayout()
        self.autosend_status = QLabel("Idle.")
        self.autosend_status.setStyleSheet("padding: 2px 4px;")
        self.autosend_status.setWordWrap(True)
        auto_row.addWidget(self.autosend_status, stretch=1)

        self.send_now_btn = QPushButton("Send Remaining Now")
        self.send_now_btn.setToolTip("Send the next batch immediately (still limited by your 24h cap).")
        self.send_now_btn.clicked.connect(self.send_remaining_now)
        auto_row.addWidget(self.send_now_btn)

        self.cancel_pending_btn = QPushButton("Cancel Pending")
        self.cancel_pending_btn.setToolTip("Stop the current send and clear the queued (paused) emails.")
        self.cancel_pending_btn.clicked.connect(self.cancel_pending)
        auto_row.addWidget(self.cancel_pending_btn)
        cred_layout.addLayout(auto_row)

        cred_box.setLayout(cred_layout)
        main_layout.addWidget(cred_box)

        self.update_sent_counter_label()

        # Timer: every 30s, refresh the countdown and auto-resume when there's
        # headroom and we're inside the allowed hours (8 AM - 9 PM).
        from PySide6.QtCore import QTimer
        self.autosend_timer = QTimer(self)
        self.autosend_timer.setInterval(30 * 1000)
        self.autosend_timer.timeout.connect(self._autosend_tick)
        self.autosend_timer.start()
        self._refresh_autosend_status()

        widget = QWidget()
        widget.setLayout(main_layout)
        self.setCentralWidget(widget)

    def save_settings(self):
        data = {
            "university_name": self.uni_input.text().strip(),
            "faculty_url": self.url_input.text().strip(),
            "campaign_folder": self.folder_input.text().strip(),
            "excel_path": self.excel_input.text().strip(),
            "gmail_address": self.gmail_input.text().strip(),
            "gmail_password": self.pass_input.text().strip(),
            "send_excel_path": self.send_excel_input.text().strip(),
            "daily_cap": self.cap_input.value(),
            "enable_ai": self.ai_checkbox.isChecked(),
            "gemini_key": self.gemini_key_input.text().strip()
        }
        try:
            with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=4)
        except Exception:
            pass

    def load_settings(self):
        if not os.path.exists(SETTINGS_FILE):
            return
        try:
            with open(SETTINGS_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)

            self.uni_input.setText(data.get("university_name", ""))
            self.url_input.setText(data.get("faculty_url", ""))
            self.folder_input.setText(data.get("campaign_folder", ""))
            self.excel_input.setText(data.get("excel_path", ""))
            self.gmail_input.setText(data.get("gmail_address", ""))
            self.pass_input.setText(data.get("gmail_password", ""))
            self.send_excel_input.setText(data.get("send_excel_path", ""))
            self.cap_input.setValue(int(data.get("daily_cap", 200) or 0))
            self.ai_checkbox.setChecked(data.get("enable_ai", False))
            self.gemini_key_input.setText(data.get("gemini_key", ""))

            folder = data.get("campaign_folder", "")
            if folder and os.path.exists(folder):
                self.load_campaign_folder(folder)
            self.update_sent_counter_label()
        except Exception as e:
            print(f"Failed to restore settings: {e}")

    def reset_settings(self):
        if os.path.exists(SETTINGS_FILE):
            os.remove(SETTINGS_FILE)
        self.uni_input.clear()
        self.url_input.clear()
        self.folder_input.clear()
        self.excel_input.clear()
        self.gmail_input.clear()
        self.pass_input.clear()
        self.send_excel_input.clear()
        self.cap_input.setValue(200)
        self.gemini_key_input.clear()
        self.ai_checkbox.setChecked(False)
        self.subject_preview.clear()
        self.body_preview.clear()
        self.campaign_data = None
        QMessageBox.information(self, "Reset Complete", "All saved settings have been cleared.")

    def browse_campaign_folder(self):
        folder_path = QFileDialog.getExistingDirectory(self, "Select Campaign Folder")
        if folder_path:
            self.folder_input.setText(folder_path)
            self.save_settings()
            self.load_campaign_folder(folder_path)

    def load_campaign_folder(self, folder_path):
        try:
            loader = FolderCampaignLoader(folder_path)
            self.campaign_data = loader.load_campaign()
            self.subject_preview.setText(self.campaign_data["subject"])
            self.body_preview.setHtml(self.campaign_data["html_body"])
        except Exception as e:
            self.subject_preview.clear()
            self.body_preview.clear()
            self.campaign_data = None
            QMessageBox.warning(self, "Folder Error", str(e))

    def browse_excel_output(self):
        file_path, _ = QFileDialog.getSaveFileName(self, "Specify Excel Output File", "", "Excel Files (*.xlsx)")
        if file_path:
            if not file_path.endswith(".xlsx"):
                file_path += ".xlsx"
            self.excel_input.setText(file_path)
            self.save_settings()

    # --- 24h send-cap + counter -------------------------------------------
    def get_daily_cap(self) -> int:
        """0 means no cap."""
        try:
            return int(self.cap_input.value())
        except Exception:
            return 0

    def _on_cap_changed(self, *_):
        self.save_settings()
        self.update_sent_counter_label()

    def update_sent_counter_label(self):
        n = _sent_last_24h(self.send_history)
        cap = self.get_daily_cap()
        if cap:
            remaining = max(0, cap - n)
            self.sent_counter_label.setText(f"Sent last 24h: {n} / {cap}  (remaining: {remaining})")
        else:
            self.sent_counter_label.setText(f"Sent last 24h: {n}  (no cap)")

    def _record_send(self, email: str):
        self.send_history.append({"email": (email or "").strip().lower(), "ts": time.time()})
        _save_send_history(self.send_history)

    # --- Send from a selected Excel file ----------------------------------
    def browse_send_excel(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Select Excel file to send from", "", "Excel Files (*.xlsx *.xlsm)")
        if file_path:
            self.send_excel_input.setText(file_path)
            self.save_settings()

    def load_targets_from_excel(self, path):
        """Read faculty rows from an .xlsx into Faculty objects. Tolerant of
        column order; matches by header name. Rows without a valid email are
        skipped."""
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        header = [(str(c).strip().lower() if c is not None else "") for c in rows[0]]

        def find_col(*names):
            for n in names:
                if n in header:
                    return header.index(n)
            return -1

        idx_email = find_col("email", "e-mail", "email address")
        idx_full = find_col("full name", "name")
        idx_first = find_col("first name", "first")
        idx_last = find_col("last name", "last")
        idx_title = find_col("title", "position")
        idx_profile = find_col("profile url", "profile", "url")

        def cell(row, i):
            if i < 0 or i >= len(row) or row[i] is None:
                return ""
            return str(row[i]).strip()

        targets = []
        for row in rows[1:]:
            email = cell(row, idx_email)
            if not email or "@" not in email or email.lower() == "missing":
                continue
            full = cell(row, idx_full)
            first = cell(row, idx_first)
            last = cell(row, idx_last)
            title = cell(row, idx_title)
            profile = cell(row, idx_profile)
            if not first and not last and full:
                parts = full.split()
                first = parts[0] if parts else ""
                last = parts[-1] if len(parts) > 1 else ""
            if not full:
                full = f"{first} {last}".strip()
            targets.append(Faculty(
                full_name=full, first_name=first, last_name=last,
                title=title, email=email.lower(), profile_url=profile,
                source_url=path, email_source="excel", bio_text=title,
            ))
        return targets

    def send_from_excel(self):
        if not self.campaign_data:
            QMessageBox.warning(self, "Error", "Please select a valid campaign folder first.")
            return
        if not self.gmail_input.text().strip() or not self.pass_input.text().strip():
            QMessageBox.warning(self, "Error", "Please enter your Gmail address and App Password.")
            return
        path = self.send_excel_input.text().strip()
        if not path or not os.path.exists(path):
            QMessageBox.warning(self, "Error", "Please choose an existing Excel (.xlsx) file to send from.")
            return
        try:
            targets = self.load_targets_from_excel(path)
        except Exception as e:
            QMessageBox.critical(self, "Excel Read Error", f"Could not read that Excel file:\n{e}")
            return
        if not targets:
            QMessageBox.warning(self, "No recipients",
                                "That Excel file had no rows with a valid email address.")
            return
        self._execute_send(targets, source_label=f"Excel file: {os.path.basename(path)}")

    def start_scan(self):
        if self._sending:
            QMessageBox.information(
                self, "Busy",
                "Emails are currently being sent. Please wait for the send to finish "
                "before starting a new scan.")
            return
        if getattr(self, "_scanning", False):
            QMessageBox.information(self, "Busy", "A scan is already running.")
            return
        url = self.url_input.text().strip()
        if not url:
            QMessageBox.warning(self, "Error", "Please enter a faculty directory URL.")
            return

        self.save_settings()
        self._scanning = True
        self._set_busy_lock()   # lock send + scan buttons while scanning
        self.status_label.setText("Scanning faculty directory...")
        self.thread = QThread()
        self.worker = ScrapeWorker(url, university_name=self.uni_input.text().strip())
        self.worker.moveToThread(self.thread)
        self.thread.started.connect(self.worker.run)
        self.worker.progress.connect(self.status_label.setText)
        self.worker.finished.connect(self.on_scan_finished)
        self.thread.start()

    def on_scan_finished(self, results):
        self.thread.quit()
        self._scanning = False
        self.faculty_data = results
        self.table.setRowCount(len(results))
        for i, fac in enumerate(results):
            self.table.setItem(i, 0, QTableWidgetItem(fac.full_name))
            self.table.setItem(i, 1, QTableWidgetItem(fac.title))
            self.table.setItem(i, 2, QTableWidgetItem(fac.email or "Missing"))
            self.table.setItem(i, 3, QTableWidgetItem("Found" if fac.email else "Needs Search"))

        self.status_label.setText(f"Scan complete. Found {len(results)} faculty records.")
        self._set_busy_lock()   # restore button states now that scanning is done

        excel_path = self.excel_input.text().strip()
        if excel_path:
            self.export_to_excel(excel_path)

    def export_to_excel(self, file_path):
        uni_name = self.uni_input.text().strip() or "Unknown University"
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Faculty Directory"

            headers = ["University", "Full Name", "First Name", "Last Name", "Title", "Email", "Profile URL", "Status"]
            ws.append(headers)

            fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            font = Font(color="FFFFFF", bold=True)
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = fill
                cell.font = font
                cell.alignment = Alignment(horizontal="center")

            for fac in self.faculty_data:
                ws.append([
                    uni_name,
                    fac.full_name,
                    fac.first_name,
                    fac.last_name,
                    fac.title,
                    fac.email or "Missing",
                    fac.profile_url,
                    "Found" if fac.email else "Missing Email"
                ])

            wb.save(file_path)
            QMessageBox.information(self, "Excel Exported", f"Saved {len(self.faculty_data)} faculty records to:\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Excel Save Error", f"Failed to save Excel file:\n{str(e)}")

    # --- Recipient filtering / safety config ---
    RESEARCH_KEYWORDS = [
        "energy", "emissions", "electric vehicle", "ev charging",
        "grid", "power grid", "electrification", "vehicle emissions",
        "transportation", "charging demand", "fuel consumption",
    ]
    MAX_RECIPIENTS_PER_RUN = 500     # Gmail's daily ceiling; sends the whole list in one run
    SEND_DELAY_SECONDS = 8           # pause between emails to look human / avoid throttling

    def send_emails(self):
        """SEND ALL EMAILS button: sends to the freshly-scraped list in the table."""
        if not self.campaign_data:
            QMessageBox.warning(self, "Error", "Please select a valid campaign folder.")
            return
        if not self.gmail_input.text().strip() or not self.pass_input.text().strip():
            QMessageBox.warning(self, "Error", "Please enter your Gmail address and App Password.")
            return

        # Build the filtered recipient list (safety role exclusions etc.).
        included, excluded = filter_recipients(
            self.faculty_data,
            keywords=self.RESEARCH_KEYWORDS,
            max_recipients=self.MAX_RECIPIENTS_PER_RUN,
            require_keyword_match=False,
        )
        if not included:
            QMessageBox.warning(
                self, "No recipients",
                "After filtering, no faculty matched your criteria.\n\n"
                "Either the crawl found no research-keyword matches, or everyone "
                "was an excluded role (president, dean, emeritus, etc.)."
            )
            return

        # Map filtered emails back to full faculty records.
        by_email = {}
        for fac in self.faculty_data:
            e = (getattr(fac, "email", "") or "").strip().lower()
            if e and e not in by_email:
                by_email[e] = fac
        targets = [by_email[row.email] for row in included if row.email in by_email]
        self._execute_send(targets, source_label="scraped list")

    def _execute_send(self, targets, source_label="list"):
        """Entry point for both send buttons. Cleans + dedupes + skips
        already-sent, queues the whole batch as 'pending', asks for one
        confirmation, then sends the first cap-limited chunk on a background
        thread. Anything over the 24h cap stays queued and auto-resumes later."""
        if self._sending:
            QMessageBox.information(self, "Already sending",
                                    "A send is already in progress. Please wait for it to finish.")
            return
        if self._scanning:
            QMessageBox.information(self, "Busy",
                                    "A faculty scan is running. Please wait for it to finish before sending.")
            return
        if not self.campaign_data:
            QMessageBox.warning(self, "Error", "Please select a valid campaign folder.")
            return
        if not self.gmail_input.text().strip() or not self.pass_input.text().strip():
            QMessageBox.warning(self, "Error", "Please enter your Gmail address and App Password.")
            return

        # 1. Clean: drop blanks, dedupe by email.
        seen, clean = set(), []
        for fac in targets:
            e = (getattr(fac, "email", "") or "").strip().lower()
            if not e or "@" not in e or e in seen:
                continue
            seen.add(e)
            clean.append(fac)

        # 2. Skip anyone already emailed in the last 24h.
        already = _emails_last_24h(self.send_history)
        skipped_already = [f for f in clean if f.email.strip().lower() in already]
        clean = [f for f in clean if f.email.strip().lower() not in already]

        if not clean:
            QMessageBox.warning(
                self, "Nothing to send",
                f"After removing blanks, duplicates, and addresses already emailed "
                f"in the last 24h, there was nothing left to send from the {source_label}.")
            return

        # 3. Work out how many fit under the cap right now.
        cap = self.get_daily_cap()
        sent24 = _sent_last_24h(self.send_history)
        headroom = (cap - sent24) if cap else len(clean)
        first_batch_n = len(clean) if headroom >= len(clean) else max(0, headroom)
        leftover_n = len(clean) - first_batch_n

        # 4. Confirmation (once).
        note = ""
        if skipped_already:
            note += f"Skipping {len(skipped_already)} already emailed in the last 24h.\n"
        if cap:
            note += f"24h cap {cap}; already sent {sent24}.\n"
        if leftover_n > 0:
            if first_batch_n > 0:
                note += (f"Will send {first_batch_n} now and automatically send the "
                         f"remaining {leftover_n} later (between "
                         f"{ALLOWED_START_HOUR%12 or 12} AM and {ALLOWED_END_HOUR%12 or 12} PM) "
                         f"as your 24h window frees up.\n")
            else:
                note += (f"Cap is already full, so all {leftover_n} will be queued and sent "
                         f"automatically once there's room.\n")
        header = (f"About to queue {len(clean)} email(s) from the {source_label}, "
                  f"{self.SEND_DELAY_SECONDS}s apart.\n{note}\n"
                  f"Scroll to review, then click Yes to proceed.")
        preview = "\n".join(f"{(f.full_name or '(no name)'):32}  <{f.email}>" for f in clean)
        if not confirm_send_dialog(self, header, preview):
            return

        # 5. Persist the WHOLE cleaned list as the pending queue (crash-safe).
        self.pending_targets = clean
        self._pending_source = source_label
        self._pending_uni = self.uni_input.text().strip()
        _save_pending_queue(self.pending_targets, self._pending_source, self._pending_uni)

        # 6. Launch the first batch (interactive: summary dialog on finish).
        self._launch_next_batch(interactive=True)

    def _launch_next_batch(self, interactive=False):
        """Send the next cap-limited slice of self.pending_targets on a worker
        thread. Leaves the rest queued for auto-resume."""
        if self._sending or self._scanning or not self.pending_targets:
            self._refresh_autosend_status()
            return
        if not self.campaign_data:
            self.autosend_status.setText("Paused: campaign folder not loaded - can't send.")
            return
        gmail = self.gmail_input.text().strip()
        pwd = self.pass_input.text().strip()
        if not gmail or not pwd:
            self.autosend_status.setText("Paused: Gmail credentials missing.")
            return

        cap = self.get_daily_cap()
        sent24 = _sent_last_24h(self.send_history)
        headroom = (cap - sent24) if cap else len(self.pending_targets)
        if headroom <= 0:
            self._refresh_autosend_status()
            return

        # Safety: never re-email anyone already sent in the last 24h (covers
        # resuming a queue after an app restart).
        already = _emails_last_24h(self.send_history)
        before = len(self.pending_targets)
        self.pending_targets = [f for f in self.pending_targets
                                if (f.email or "").strip().lower() not in already]
        if len(self.pending_targets) != before:
            _save_pending_queue(self.pending_targets, self._pending_source, self._pending_uni)
        if not self.pending_targets:
            self._refresh_autosend_status()
            return

        batch = self.pending_targets[:headroom]

        ai_personalizer = None
        if self.ai_checkbox.isChecked():
            gemini_key = self.gemini_key_input.text().strip()
            if not gemini_key:
                self.autosend_status.setText("Paused: AI is on but no Gemini key entered.")
                return
            ai_personalizer = AIPersonalizer(gemini_key)

        temp_dir = os.path.join(self.folder_input.text().strip() or ".", "temp_cvs")
        os.makedirs(temp_dir, exist_ok=True)

        self._sending = True
        self._run_failures = []
        self._run_interactive = interactive
        self._run_source = self._pending_source

        self._send_thread = QThread()
        self._send_worker = SendWorker(
            batch, self.campaign_data, ai_personalizer,
            self.uni_input.text().strip(), temp_dir, self.SEND_DELAY_SECONDS, gmail, pwd)
        self._send_worker.moveToThread(self._send_thread)
        self._send_thread.started.connect(self._send_worker.run)
        self._send_worker.progress.connect(self.autosend_status.setText)
        self._send_worker.sent_one.connect(self._on_sent_one)
        self._send_worker.failed_one.connect(self._on_failed_one)
        self._send_worker.finished.connect(self._on_send_finished)
        self._set_send_buttons_enabled(False)
        self._send_thread.start()

    def _on_sent_one(self, email):
        self._record_send(email)
        # remove from pending queue (crash-safe shrink)
        el = email.strip().lower()
        self.pending_targets = [f for f in self.pending_targets
                                if (f.email or "").strip().lower() != el]
        _save_pending_queue(self.pending_targets, self._pending_source, self._pending_uni)
        self.update_sent_counter_label()

    def _on_failed_one(self, email, err):
        self._run_failures.append((email, err))
        # drop failed address from the queue too, so it isn't retried forever
        el = email.strip().lower()
        self.pending_targets = [f for f in self.pending_targets
                                if (f.email or "").strip().lower() != el]
        _save_pending_queue(self.pending_targets, self._pending_source, self._pending_uni)

    def _on_send_finished(self, sent_count, fail_count):
        if self._send_thread:
            self._send_thread.quit()
            self._send_thread.wait(2000)
        self._sending = False
        self._send_thread = None
        self._send_worker = None
        self._set_send_buttons_enabled(True)
        self.update_sent_counter_label()

        remaining = len(self.pending_targets)
        if self._run_interactive:
            msg = f"Sent {sent_count} email(s) from the {getattr(self, '_run_source', 'list')}."
            if fail_count:
                msg += f"\n{fail_count} failed. First few:\n"
                for email, err in self._run_failures[:5]:
                    msg += f"  {email}: {err}\n"
            msg += f"\nSent in last 24h now: {_sent_last_24h(self.send_history)}"
            cap = self.get_daily_cap()
            if cap:
                msg += f" / cap {cap}."
            if remaining:
                dt = _effective_resume_dt(self.send_history, cap, min(remaining, cap or remaining))
                msg += (f"\n\n{remaining} still queued (over today's cap). "
                        f"They'll auto-send starting ~{_format_clock(dt)}.")
            QMessageBox.information(self, "Email Dispatch Complete", msg)

        self._refresh_autosend_status()

    def _set_busy_lock(self):
        """Single source of truth for which buttons are clickable. While a send
        or a scan is running, the SEND buttons and the crawl button are locked so
        you can't start two operations at once (which is what froze the app)."""
        busy = self._sending or self._scanning
        # Send controls: locked whenever anything is running.
        for btn in (getattr(self, "send_btn", None),
                    getattr(self, "send_from_excel_btn", None),
                    getattr(self, "send_now_btn", None)):
            if btn is not None:
                btn.setEnabled(not busy)
        # Crawl button: locked whenever anything is running.
        if getattr(self, "scan_btn", None) is not None:
            self.scan_btn.setEnabled(not busy)

    def _set_send_buttons_enabled(self, enabled):
        # kept for compatibility with existing calls; delegates to the lock
        self._set_busy_lock()

    # --- Auto-resume timer + status ---------------------------------------
    def _autosend_tick(self):
        if self._sending or self._scanning:
            return
        if not self.pending_targets:
            self._refresh_autosend_status()
            return
        cap = self.get_daily_cap()
        sent24 = _sent_last_24h(self.send_history)
        headroom = (cap - sent24) if cap else len(self.pending_targets)
        remaining = len(self.pending_targets)
        need = min(remaining, cap) if cap else remaining
        # Auto-resume only inside allowed hours, and only when there's room for
        # the remaining batch (or a full cap's worth if remaining exceeds cap).
        if _in_allowed_hours() and headroom >= need and self.campaign_data \
           and self.gmail_input.text().strip() and self.pass_input.text().strip():
            self.autosend_status.setText(f"Auto-resuming: sending {min(headroom, remaining)} queued email(s)...")
            self._launch_next_batch(interactive=False)
        else:
            self._refresh_autosend_status()

    def _refresh_autosend_status(self):
        remaining = len(self.pending_targets)
        if self._sending:
            return  # progress text is driven by the worker
        if not remaining:
            self.autosend_status.setText("Idle. No emails queued.")
            return
        cap = self.get_daily_cap()
        need = min(remaining, cap) if cap else remaining
        dt = _effective_resume_dt(self.send_history, cap, need)
        now = datetime.now()
        if dt <= now and _in_allowed_hours(now):
            self.autosend_status.setText(
                f"{remaining} queued - resuming shortly...")
        else:
            self.autosend_status.setText(
                f"Paused: {remaining} email(s) queued. Auto-resumes ~{_format_clock(dt)}.")

    def send_remaining_now(self):
        """Force the next cap-limited batch immediately (ignores the wait, but
        still respects the 24h cap so you can't overshoot)."""
        if not self.pending_targets:
            QMessageBox.information(self, "Nothing queued", "There are no queued emails to send.")
            return
        cap = self.get_daily_cap()
        sent24 = _sent_last_24h(self.send_history)
        if cap and sent24 >= cap:
            dt = _effective_resume_dt(self.send_history, cap, min(len(self.pending_targets), cap))
            QMessageBox.warning(
                self, "Daily cap reached",
                f"You've sent {sent24}/{cap} in the last 24h. There's no room right now.\n"
                f"Room opens ~{_format_clock(dt)}, or raise the cap.")
            return
        self._launch_next_batch(interactive=True)

    def cancel_pending(self):
        if not self.pending_targets and not self._sending:
            QMessageBox.information(self, "Nothing to cancel", "There are no queued or in-progress emails.")
            return
        confirm = QMessageBox.question(
            self, "Cancel pending send",
            "Stop the current send (after the current email) and clear all queued emails?")
        if confirm != QMessageBox.Yes:
            return
        if self._sending and self._send_worker:
            self._send_worker.stop()
        self.pending_targets = []
        _save_pending_queue([], "", "")
        self.autosend_status.setText("Cancelled. Queue cleared.")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())